# -*- coding: utf-8 -*-
"""
Created on Thu Feb 10 16:04:55 2022

@author: Tanner Law
"""
#imports

import sys
import logging
from openpyxl import load_workbook


#global vars
wb = ""



#functions
def main():
    #create logger
    myLog = logger()
    workbook = load_workbook("expedia_report_monthly_march_2018.xlsx")
    
    try:
        #if they don't pass in what they want assume this file
        if(len(sys.argv)<2):
            path = input("Please enter in the file location of the expedia report you want: ")
            workbook = load_workbook(path)
            #otherwise take what was passed in.
        else:
            workbook = load_workbook(sys.argv[1])
            wb = sys.argv[1]
    except:
        myLog.error("Could not open workbook.")
    
    #print(workbook.sheetnames)
    workbook.move_sheet("Summary Rolling MoM")
    sheet = workbook.active

    
    getData( sheet, myLog, path)

    
    
#create logger
def logger():
    #Creating and Configuring Logger

    Log_Format = "%(asctime)s - %(message)s"

    logging.basicConfig(filename = "logfile.log",
                    filemode = "w",
                    format = Log_Format, 
                    level = logging.ERROR)

    logger = logging.getLogger()
    #logger.error("Our First Log Message")
    
    return logger

#find what row I want
def wantedRow(myLog, wb, sheet):
    #get the month and year to loop through on the file
    wb = wb.split("\\")[-1]
    #print(wb)
    wb = wb.split('.')[0]

    #print(wb)
    month = wb.split('_')[-2]
    year = wb.split('_')[-1]
    year = year[2] + year[3]
    #print(year)
    #print(month)
    #print(rowWanted)
    #print(monthNum(month))
    
    #find the row containing the str I gor for rowWanted
    myLog.error("getting row num I need")
    

    for x in range(2, 50):
        #print(sheet.cell(row=x, column=1).value)
        val = str(sheet.cell(row=x, column=1).value)
        val = val.split(" ")[0]
        if(val.split("-")[-1] == year and int(val.split("-")[-2]) == monthNum(month)):
            #print("row I want is: " + str(x))
            return x

#get my data from a given row    
def getData(sheet, myLog, path):
    myLog.error("Getting Data")
    
    valuesDict = ("Calls Offered", "Abandoned after 30s", "FCR", "DSAT", "CSAT")
    values = []
    
    myRow = wantedRow(myLog, path, sheet)
    
    #print("my row is: {}".format(myRow))
    
    #This isn't working fix me *****************************************************************************
    #for i in range(2,7):
    #    values.append(sheet.cell(row = myRow, column = i).value)
    values.append(sheet.cell(row = myRow, column = range(2,7)).value)
        
        
    myLog.error("Data for {}: \n".format(str(sheet.cell(row = myRow, column = 1).value).split(" ")[0]))
        
    for i in range(0, len(valuesDict)):
        if i == 0:
            
            myLog.error("{}: {} ".format(valuesDict[i], values[i]))
        else:
            myLog.error("{} : {} : {}%".format(valuesDict[i], values[i], (values[i]*100)))
    
#get month num
def monthNum(str):
    months = {"january": 1,
              "february": 2,
              "march": 3,
              "april": 4,
              "may": 5,
              "june": 6,
              "july": 7,
              "august": 8,
              "september": 9,
              "october": 10,
              "november": 11,
              "december": 12,
              }
    return months[str]

#start the program by calling the main function
main()